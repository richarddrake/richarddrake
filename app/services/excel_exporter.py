# 这个模块负责把生成后的测试用例整理成 Excel 文件并应用统一表格样式。
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas import TestCase


HEADERS = [
    ("id", "用例编号"),
    ("module", "模块"),
    ("title", "用例标题"),
    ("priority", "优先级"),
    ("case_type", "类型"),
    ("scenario", "覆盖场景"),
    ("preconditions", "前置条件"),
    ("steps", "操作步骤"),
    ("expected_results", "预期结果"),
    ("test_data", "测试数据"),
    ("tags", "标签"),
    ("source", "依据"),
    ("quality", "质量评分"),
    ("coverage", "覆盖分析"),
    ("api_test", "可执行接口配置"),
    ("ui_test", "可执行 UI 配置"),
    ("execution", "最近执行结果"),
]


def save_cases_to_excel(cases: list[TestCase], target_dir: Path, session_id: str) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试用例"

    sheet.append([label for _, label in HEADERS])

    for case in cases:
        item = case.to_dict()
        sheet.append([_cell_value(item.get(key)) for key, _ in HEADERS])

    _style_sheet(sheet)

    path = target_dir / f"test_cases_{session_id}.xlsx"
    workbook.save(path)
    return path


def _cell_value(value: object) -> str:
    if isinstance(value, list):
        return "\n".join(f"{idx}. {item}" for idx, item in enumerate(value, 1))
    if isinstance(value, dict):
        if "score" in value:
            issues = value.get("issues") or []
            suggestions = value.get("suggestions") or []
            return "\n".join(
                [
                    f"score: {value.get('score')}",
                    f"level: {value.get('level', '')}",
                    f"issues: {'；'.join(str(item) for item in issues)}",
                    f"suggestions: {'；'.join(str(item) for item in suggestions)}",
                ]
            ).strip()
        return json.dumps(value, ensure_ascii=False, indent=2)
    if value is None:
        return ""
    return str(value)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="E8F0FE")
    header_font = Font(bold=True, color="1F1F1F")
    thin = Side(style="thin", color="DADCE0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    widths = [14, 18, 34, 10, 12, 30, 34, 42, 42, 24, 20, 28, 34, 34, 48, 34]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
